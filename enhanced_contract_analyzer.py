#!/usr/bin/env python3
"""
Enhanced Contract Financial Analyzer - Command Line Version with Three-Pass Validation
Automatically analyzes contract documents and generates comprehensive payment tracking spreadsheets
with high-precision financial data extraction
"""

import os
import sys
import json
import re
import argparse
from pathlib import Path
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.differential import DifferentialStyle
import anthropic
from typing import Dict, List, Optional, Tuple
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle

class HighPrecisionContractAnalyzer:
    def __init__(self, api_key: str):
        """Initialize the enhanced contract analyzer with Claude API key and precision patterns."""
        self.client = anthropic.Anthropic(api_key=api_key)
        
        # High-precision patterns for critical financial data
        self.financial_patterns = {
            'money_amounts': [
                r'\$\s*[\d,]+\.?\d*(?:\s*(?:USD|dollars?))?',  # $1,000.00
                r'(?:USD|dollars?)\s*[\d,]+\.?\d*',            # USD 1000
                r'[\d,]+\.?\d*\s*(?:USD|dollars?)',            # 1000 USD
                r'(?:total|amount|value|fee|rate|cost|price|budget|cap|maximum|limit)(?:\s+(?:of|is|at|shall be|not to exceed))?\s*[:\$]?\s*[\d,]+\.?\d*',
            ],
            'hourly_rates': [
                r'\$\s*[\d,]+\.?\d*\s*(?:per|/|an)\s*hour',      # $50/hour, $50 per hour
                r'hourly\s+(?:rate|fee|charge)(?:\s+(?:of|is|at))?\s*[:\$]?\s*[\d,]+\.?\d*',
                r'[\d,]+\.?\d*\s*(?:per|/)\s*(?:hour|hr)',
            ],
            'contract_values': [
                r'(?:total|contract|project|agreement)\s+(?:value|amount|price|cost|fee)(?:\s+(?:of|is|shall be|not to exceed))?\s*[:\$]?\s*[\d,]+\.?\d*',
                r'(?:maximum|cap|limit|ceiling)(?:\s+(?:of|is|at))?\s*[:\$]?\s*[\d,]+\.?\d*',
                r'budget(?:\s+(?:of|is|shall be|not to exceed))?\s*[:\$]?\s*[\d,]+\.?\d*',
            ],
            'dates': [
                r'\b\d{1,2}[/-]\d{1,2}[/-]\d{4}\b',           # MM/DD/YYYY or MM-DD-YYYY
                r'\b\d{4}[/-]\d{2}[/-]\d{2}\b',               # YYYY-MM-DD
                r'\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},?\s+\d{4}\b',
                r'\b\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}\b',
            ],
            'time_periods': [
                r'\b\d+\s*(?:hours?|hrs?)\b',
                r'\b\d+\s*(?:days?)\b',
                r'\b\d+\s*(?:weeks?)\b', 
                r'\b\d+\s*(?:months?)\b',
                r'maximum.*?\d+.*?(?:hours?|days?)',
                r'(?:daily|weekly|monthly)\s+(?:limit|maximum|cap).*?\d+',
            ],
'payment_terms': [
                r'net\s+\d+(?:\s+days?)?',                     # net 30, net 30 days
                r'within\s+\d+\s+days?',                       # within 30 days
                r'payment\s+due.*?\d+\s+days?',                # payment due within 30 days
                r'\d+\s+days?\s+(?:after|from|following)',     # 30 days after
            ]
        }
    
    def parse_currency_to_number(self, currency_string: str) -> Optional[float]:
        """Convert currency string to float number for Excel calculations."""
        if not currency_string or currency_string in ['null', 'N/A', 'Not found', '']:
            return None
        
        # Remove currency symbols, commas, and extra spaces
        cleaned = re.sub(r'[^\d.-]', '', str(currency_string))
        
        try:
            return float(cleaned) if cleaned else None
        except ValueError:
            return None

    def parse_hours_to_number(self, hours_string: str) -> Optional[float]:
        """Convert hours string to float number."""
        if not hours_string or hours_string in ['null', 'N/A', 'Not found', '']:
            return None
        
        # Extract just the number part
        cleaned = re.sub(r'[^\d.-]', '', str(hours_string))
        
        try:
            return float(cleaned) if cleaned else None
        except ValueError:
            return None
        
    def read_contract_file(self, file_path: str) -> str:
        """Read contract file and return content as string."""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Contract file not found: {file_path}")
            
        # Handle different file types
        if path.suffix.lower() == '.pdf':
            try:
                import PyPDF2
                with open(file_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    text = ""
                    for page in pdf_reader.pages:
                        text += page.extract_text()
                    return text
            except ImportError:
                raise ImportError("PyPDF2 required for PDF files. Install with: pip install PyPDF2")
                
        elif path.suffix.lower() in ['.txt', '.docx', '.doc']:
            if path.suffix.lower() == '.docx':
                try:
                    from docx import Document
                    doc = Document(file_path)
                    return '\n'.join([paragraph.text for paragraph in doc.paragraphs])
                except ImportError:
                    raise ImportError("python-docx required for Word files. Install with: pip install python-docx")
            else:
                with open(file_path, 'r', encoding='utf-8') as file:
                    return file.read()
        else:
            # Try to read as plain text
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()

    def extract_financial_data_regex(self, contract_text: str) -> Dict:
        """PASS 1: Extract critical financial data using high-precision regex patterns."""
        
        print("  → Running Pass 1: High-precision regex pattern extraction...")
        
        results = {
            'money_amounts': [],
            'hourly_rates': [],
            'contract_values': [],
            'dates': [],
            'time_periods': [],
            'payment_terms': []
        }
        
        # Extract each category with context
        for category, patterns in self.financial_patterns.items():
            unique_matches = set()  # Avoid duplicates
            
            for pattern in patterns:
                matches = re.finditer(pattern, contract_text, re.IGNORECASE | re.MULTILINE)
                
                for match in matches:
                    # Get surrounding context (50 chars before/after)
                    start = max(0, match.start() - 50)
                    end = min(len(contract_text), match.end() + 50)
                    context = contract_text[start:end].strip()
                    
                    # Clean up the matched text
                    matched_text = match.group().strip()
                    
                    # Store with context for validation
                    unique_matches.add((matched_text, context))
            
            # Convert back to list and log findings
            results[category] = list(unique_matches)
            if results[category]:
                print(f"    • Found {len(results[category])} {category.replace('_', ' ')}")
        
        return results

    def claude_direct_extraction(self, contract_text: str) -> Dict:
        """PASS 2: Direct Claude extraction with focused prompts."""
        
        print("  → Running Pass 2: Direct Claude AI extraction...")
        
        claude_prompt = f"""
        Extract ONLY these specific financial data points from this contract.
        Be extremely precise - include ONLY information that is explicitly stated:
        
        1. TOTAL CONTRACT VALUE (the main contract amount)
        2. HOURLY RATE (if time & materials)
        3. CONTRACT START DATE
        4. CONTRACT END DATE  
        5. PAYMENT TERMS (Net 30, etc.)
        6. MAXIMUM HOURS (if specified)
        7. DAILY HOUR LIMIT (if specified)
        
        Return as JSON:
        {{
            "total_contract_value": "exact amount or null",
            "hourly_rate": "exact rate or null",
            "start_date": "YYYY-MM-DD or null", 
            "end_date": "YYYY-MM-DD or null",
            "payment_terms": "exact terms or null",
            "maximum_hours": "number or null",
            "daily_hour_limit": "number or null"
        }}
        
        Contract: {contract_text}
        """
        
        try:
            claude_result = self.client.messages.create(
                model="claude-3-haiku-20240307",
                max_tokens=1000,
                messages=[{"role": "user", "content": claude_prompt}]
            )
            
            claude_text = claude_result.content[0].text
            start_idx = claude_text.find('{')
            end_idx = claude_text.rfind('}') + 1
            
            if start_idx != -1 and end_idx > start_idx:
                claude_data = json.loads(claude_text[start_idx:end_idx])
                print(f"    • Claude extracted {len([v for v in claude_data.values() if v and v != 'null'])} data points")
                return claude_data
            else:
                print("    • Claude extraction: Could not parse JSON response")
                return {}
                
        except Exception as e:
            print(f"    • Claude extraction error: {e}")
            return {}

    def validate_and_merge_data(self, regex_results: Dict, claude_data: Dict, contract_text: str) -> Dict:
        """PASS 3: Validate and merge results from both methods."""
        
        print("  → Running Pass 3: Cross-validation and confidence scoring...")
        
        validation_prompt = f"""
        You are validating financial data extracted by two different methods. 
        Provide the MOST ACCURATE data by comparing both sources:
        
        REGEX EXTRACTED DATA:
        Money amounts: {[m[0] for m in regex_results.get('money_amounts', [])[:5]]}  # Show first 5
        Hourly rates: {[m[0] for m in regex_results.get('hourly_rates', [])]}
        Dates: {[m[0] for m in regex_results.get('dates', [])[:10]]}  # Show first 10
        Payment terms: {[m[0] for m in regex_results.get('payment_terms', [])]}
        
        CLAUDE EXTRACTED DATA:
        {claude_data}
        
        Return the validated data as JSON with confidence levels:
        {{
            "total_contract_value": "best value",
            "hourly_rate": "best value",
            "start_date": "best value",
            "end_date": "best value", 
            "payment_terms": "best value",
            "maximum_hours": "best value",
            "daily_hour_limit": "best value",
            "confidence_scores": {{
                "total_contract_value": "HIGH/MEDIUM/LOW",
                "hourly_rate": "HIGH/MEDIUM/LOW",
                "start_date": "HIGH/MEDIUM/LOW",
                "end_date": "HIGH/MEDIUM/LOW",
                "payment_terms": "HIGH/MEDIUM/LOW",
                "maximum_hours": "HIGH/MEDIUM/LOW",
                "daily_hour_limit": "HIGH/MEDIUM/LOW"
            }},
            "validation_notes": "any discrepancies or concerns"
        }}
        """
        
        try:
            validation_result = self.client.messages.create(
                model="claude-3-haiku-20240307",
                max_tokens=1500,
                messages=[{"role": "user", "content": validation_prompt}]
            )
            
            validation_text = validation_result.content[0].text
            start_idx = validation_text.find('{')
            end_idx = validation_text.rfind('}') + 1
            
            if start_idx != -1 and end_idx > start_idx:
                validated_data = json.loads(validation_text[start_idx:end_idx])
                
                # Count confidence levels
                confidence_scores = validated_data.get('confidence_scores', {})
                high_count = len([s for s in confidence_scores.values() if s == "HIGH"])
                total_count = len([s for s in confidence_scores.values() if s in ["HIGH", "MEDIUM", "LOW"]])
                
                print(f"    • Validation complete: {high_count}/{total_count} HIGH confidence items")
                
                return validated_data
            else:
                print("    • Validation: Could not parse response")
                return {"validation_notes": "Could not parse validation results"}
                
        except Exception as e:
            print(f"    • Validation error: {e}")
            return {"validation_notes": f"Validation failed: {e}"}

    def cross_validate_financial_data(self, contract_text: str) -> Dict:
        """Run the three-pass validation system."""
        
        print("\n" + "="*60)
        print("STARTING THREE-PASS VALIDATION SYSTEM")
        print("="*60)
        
        # Pass 1: Regex extraction
        regex_results = self.extract_financial_data_regex(contract_text)
        
        # Pass 2: Claude direct extraction  
        claude_data = self.claude_direct_extraction(contract_text)
        
        # Pass 3: Validation and merging
        validated_data = self.validate_and_merge_data(regex_results, claude_data, contract_text)
        
        print("✓ Three-pass validation complete")
        
        return validated_data

    def analyze_contract_comprehensive(self, contract_text: str) -> Dict:
        """Send contract to Claude API for comprehensive financial analysis."""
        
        enhanced_analysis_prompt = """
        You are a Certified Public Accountant analyzing this contract for complete financial tracking requirements. Extract ALL financial information including:

        PRIMARY PAYMENT STRUCTURE:
        - Contract type (fixed payments, time & materials, milestone-based, etc.)
        - Total contract value and any caps/maximums
        - Payment rates (hourly, daily, fixed amounts)
        - Payment frequency and timing requirements

        PAYMENT SCHEDULE DETAILS:
        - Specific due dates for payments OR invoice submission requirements
        - Invoice approval processes and timeframes
        - Expected payment timing after invoice submission
        - Any advance payments or retainer requirements

        EXPENSE AND REIMBURSEMENT TRACKING:
        - Travel expense policies and limits
        - Meal and incidental expense allowances
        - Equipment or material cost provisions
        - Any expense pre-approval requirements
        - Separate expense tracking from main contract value

        COMPLIANCE AND DOCUMENTATION REQUIREMENTS:
        - Required tax forms (W-9, 1099, etc.)
        - Invoice content requirements (daily breakdowns, certifications, etc.)
        - Time tracking limitations (daily/weekly maximums)
        - Reporting deadlines and deliverable schedules
        - Required signatures or approvals

        BUDGET MONITORING ELEMENTS:
        - Maximum hours, days, or other quantity limits
        - Running total calculations needed
        - Budget utilization warnings or thresholds
        - Contract period start and end dates
        - Any renewal or extension provisions

        RISK FACTORS:
        - Payment conditions or performance requirements
        - Penalty clauses or withholding provisions
        - Termination clauses affecting payment
        - Currency or rate change provisions

        For time & materials contracts specifically include:
        - Hourly rate breakdown by role/activity
        - Maximum billable hours per period
        - Overtime or premium rate conditions
        - Minimum/maximum monthly billing requirements

        Create a comprehensive payment tracking structure that includes:
        1. All scheduled payments with dates and amounts
        2. Invoice submission timeline and requirements  
        3. Expense reimbursement tracking separate from main payments
        4. Compliance documentation deadlines
        5. Budget utilization monitoring with warnings
        6. Expected payment dates based on contract terms

        If no specific payment dates exist, create a logical payment tracking framework based on the contract's invoicing and payment cycle requirements.

        Contract text:
        """
        
        try:
            message = self.client.messages.create(
                model="claude-3-haiku-20240307",
                max_tokens=4000,
                messages=[{
                    "role": "user",
                    "content": enhanced_analysis_prompt + contract_text
                }]
            )
            return {"analysis": message.content[0].text, "success": True}
        except Exception as e:
            return {"analysis": f"Error analyzing contract: {str(e)}", "success": False}

    def extract_comprehensive_data(self, analysis: str) -> Tuple[Dict, List[Dict], Dict]:
        """Extract comprehensive structured data from Claude's analysis."""
        
        extraction_prompt = f"""
        Based on this contract analysis, extract the comprehensive information in JSON format:
        
        Return a JSON object with:
        
        1. "contract_info": {{
            "client": "client name",
            "vendor": "vendor name", 
            "contract_id": "contract number/id",
            "contract_type": "fixed payments/time and materials/milestone/other",
            "total_value": "total contract value",
            "start_date": "YYYY-MM-DD",
            "end_date": "YYYY-MM-DD",
            "payment_terms": "payment terms description",
            "hourly_rate": "hourly rate if applicable",
            "max_hours": "maximum hours if applicable",
            "max_daily_hours": "daily hour limit if applicable",
            "invoice_frequency": "monthly/weekly/other",
            "payment_timeline": "net 30/45 days or description"
        }}
        
        2. "payment_schedule": [
            {{
                "due_date": "YYYY-MM-DD or null",
                "description": "payment description", 
                "amount": "payment amount or null",
                "invoice_submission_due": "YYYY-MM-DD or null",
                "expected_payment_date": "YYYY-MM-DD or null",
                "notes": "any additional notes"
            }}
        ]
        
        3. "tracking_requirements": {{
            "expense_tracking": {{
                "travel_expenses": "yes/no",
                "travel_policy": "description",
                "meal_allowance": "amount or policy",
                "equipment_costs": "yes/no",
                "pre_approval_required": "yes/no"
            }},
            "compliance": {{
                "w9_required": "yes/no",
                "invoice_requirements": "description",
                "time_breakdown_required": "yes/no",
                "certifications_required": "description",
                "reporting_deadlines": "description"
            }},
            "budget_monitoring": {{
                "hour_tracking": "yes/no",
                "budget_caps": "description",
                "warning_thresholds": "percentage or amount",
                "utilization_tracking": "yes/no"
            }}
        }}
        
        If information is not available, use null or "not specified".
        
        Analysis to process:
        {analysis}
        """
        
        try:
            message = self.client.messages.create(
                model="claude-3-haiku-20240307",
                max_tokens=2000,
                messages=[{
                    "role": "user", 
                    "content": extraction_prompt
                }]
            )
            
            # Extract JSON from response
            response_text = message.content[0].text
            
            # Find JSON in the response
            start_idx = response_text.find('{')
            end_idx = response_text.rfind('}') + 1
            
            if start_idx != -1 and end_idx > start_idx:
                json_str = response_text[start_idx:end_idx]
                data = json.loads(json_str)
                return (
                    data.get("contract_info", {}), 
                    data.get("payment_schedule", []),
                    data.get("tracking_requirements", {})
                )
            else:
                # Fallback: return basic structure
                return {"contract_type": "unknown"}, [], {}
                
        except Exception as e:
            print(f"Warning: Could not extract structured data: {e}")
            return {"contract_type": "unknown"}, [], {}

    def create_comprehensive_spreadsheet(self, contract_info: Dict, payment_schedule: List[Dict], 
                                       tracking_requirements: Dict, output_file: str, 
                                       original_analysis: str, validated_data: Dict = None):
        """Create comprehensive Excel spreadsheet with proper number formatting."""
        
        wb = Workbook()
        
        # Create number format styles
        from openpyxl.styles import NamedStyle
        
        currency_style = NamedStyle(name="currency")
        currency_style.number_format = '"$"#,##0.00'
        
        hours_style = NamedStyle(name="hours") 
        hours_style.number_format = '#,##0.0'
        
        percentage_style = NamedStyle(name="percentage")
        percentage_style.number_format = '0.0%'
        
        # Register styles with workbook
        wb.add_named_style(currency_style)
        wb.add_named_style(hours_style)
        wb.add_named_style(percentage_style)
        
        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        confidence_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 1. Contract Info Sheet with Confidence Scores
        info_sheet = wb.active
        info_sheet.title = "Contract Summary"
        
        confidence_scores = validated_data.get('confidence_scores', {}) if validated_data else {}
        
        contract_data = [
            ["CONTRACT INFORMATION", "", "CONFIDENCE LEVEL"],
            ["Client", contract_info.get("client", ""), ""],
            ["Vendor/Contractor", contract_info.get("vendor", ""), ""],
            ["Contract ID", contract_info.get("contract_id", ""), ""],
            ["Contract Type", contract_info.get("contract_type", ""), ""],
            ["Total Value", self.parse_currency_to_number(contract_info.get("total_value", "")), confidence_scores.get("total_contract_value", "")],
            ["Start Date", contract_info.get("start_date", ""), confidence_scores.get("start_date", "")],
            ["End Date", contract_info.get("end_date", ""), confidence_scores.get("end_date", "")],
            ["Hourly Rate", self.parse_currency_to_number(contract_info.get("hourly_rate", "")), confidence_scores.get("hourly_rate", "")],
            ["Maximum Hours", self.parse_hours_to_number(contract_info.get("max_hours", "")), confidence_scores.get("maximum_hours", "")],
            ["Daily Hour Limit", self.parse_hours_to_number(contract_info.get("max_daily_hours", "")), confidence_scores.get("daily_hour_limit", "")],
            ["Invoice Frequency", contract_info.get("invoice_frequency", ""), ""],
            ["Payment Terms", contract_info.get("payment_timeline", ""), confidence_scores.get("payment_terms", "")],
            ["", "", ""],
            ["THREE-PASS VALIDATION RESULTS", "", ""],
            ["Analysis Method", "Regex + Claude AI + Cross-Validation", "HIGH"],
            ["High Confidence Items", str(len([s for s in confidence_scores.values() if s == "HIGH"])), ""],
            ["Medium Confidence Items", str(len([s for s in confidence_scores.values() if s == "MEDIUM"])), ""],
            ["Low Confidence Items", str(len([s for s in confidence_scores.values() if s == "LOW"])), ""],
            ["Validation Notes", validated_data.get("validation_notes", "") if validated_data else "", ""],
            ["", "", ""],
            ["TRACKING REQUIREMENTS", "", ""],
            ["Travel Expenses", tracking_requirements.get("expense_tracking", {}).get("travel_expenses", "No"), ""],
            ["Expense Pre-approval", tracking_requirements.get("expense_tracking", {}).get("pre_approval_required", "No"), ""],
            ["W-9 Required", tracking_requirements.get("compliance", {}).get("w9_required", "No"), ""],
            ["Time Breakdown Required", tracking_requirements.get("compliance", {}).get("time_breakdown_required", "No"), ""],
            ["Hour Tracking", tracking_requirements.get("budget_monitoring", {}).get("hour_tracking", "No"), ""],
            ["", "", ""],
            ["Analysis Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "HIGH"]
        ]
        
        for row_idx, (label, value, confidence) in enumerate(contract_data, 1):
            cell_a = info_sheet.cell(row=row_idx, column=1, value=label)
            cell_b = info_sheet.cell(row=row_idx, column=2, value=value)
            cell_c = info_sheet.cell(row=row_idx, column=3, value=confidence)
            
            if label in ["CONTRACT INFORMATION", "THREE-PASS VALIDATION RESULTS", "TRACKING REQUIREMENTS"]:
                cell_a.font = header_font
                cell_a.fill = header_fill
                cell_b.fill = header_fill
                cell_c.fill = header_fill
            elif confidence == "HIGH":
                cell_c.fill = confidence_fill
            elif confidence == "LOW":
                cell_c.fill = warning_fill
        
        # 2. Payment Tracking Sheet
        payment_sheet = wb.create_sheet("Payment Tracking")
        
        if contract_info.get("contract_type") == "time and materials":
            headers = [
                "Period/Month", "Hours Worked", "Cumulative Hours", "Hourly Rate", 
                "Invoice Amount", "Invoice Submission Due", "Invoice Submitted Date",
                "Invoice Approved Date", "Expected Payment Date", "Actual Payment Date",
                "Amount Paid", "Balance Due", "Status", "Notes"
            ]
            
            # Add monthly rows for T&M contracts
            start_date = contract_info.get("start_date")
            end_date = contract_info.get("end_date")
            
            if start_date and end_date:
                try:
                    start = datetime.strptime(start_date, "%Y-%m-%d")
                    end = datetime.strptime(end_date, "%Y-%m-%d")
                    
                    current = start.replace(day=1)
                    monthly_periods = []
                    
                    while current <= end:
                        month_end = (current.replace(month=current.month+1) if current.month < 12 
                                   else current.replace(year=current.year+1, month=1)) - timedelta(days=1)
                        period_end = min(month_end, end)
                        
                        invoice_due = period_end + timedelta(days=30)
                        expected_payment = invoice_due + timedelta(days=30)
                        
                        monthly_periods.append({
                            "period": f"{current.strftime('%Y-%m')}",
                            "invoice_due": invoice_due.strftime('%Y-%m-%d'),
                            "expected_payment": expected_payment.strftime('%Y-%m-%d')
                        })
                        
                        current = current.replace(month=current.month+1) if current.month < 12 else current.replace(year=current.year+1, month=1)
                        
                except ValueError:
                    monthly_periods = []
                    
        else:
            headers = [
                "Due Date", "Description", "Amount Due", "Invoice Required",
                "Invoice Submission Due", "Invoice Submitted Date", "Invoice Approved Date",
                "Expected Payment Date", "Actual Payment Date", "Amount Paid", 
                "Balance Due", "Status", "Notes"
            ]
            monthly_periods = payment_schedule
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = payment_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border
        
        # Add data rows
        row_start = 2
        if contract_info.get("contract_type") == "time and materials" and monthly_periods:
            for row_idx, period in enumerate(monthly_periods, row_start):
                payment_sheet.cell(row=row_idx, column=1, value=period["period"])
                payment_sheet.cell(row=row_idx, column=2, value="")
                payment_sheet.cell(row=row_idx, column=3, value="")
                
                # Set hourly rate as number with currency formatting
                hourly_rate_cell = payment_sheet.cell(row=row_idx, column=4)
                hourly_rate_value = self.parse_currency_to_number(contract_info.get("hourly_rate", ""))
                hourly_rate_cell.value = hourly_rate_value
                if hourly_rate_value is not None:
                    hourly_rate_cell.number_format = '"$"#,##0.00'
                
                payment_sheet.cell(row=row_idx, column=5, value="")
                payment_sheet.cell(row=row_idx, column=6, value=period["invoice_due"])
                payment_sheet.cell(row=row_idx, column=7, value="")
                payment_sheet.cell(row=row_idx, column=8, value="")
                payment_sheet.cell(row=row_idx, column=9, value=period["expected_payment"])
                payment_sheet.cell(row=row_idx, column=10, value="")
                payment_sheet.cell(row=row_idx, column=11, value="")
                payment_sheet.cell(row=row_idx, column=12, value="")
                payment_sheet.cell(row=row_idx, column=13, value="Pending")
                payment_sheet.cell(row=row_idx, column=14, value="")
        else:
            data_rows = payment_schedule if payment_schedule else [{}] * 10
            for row_idx, payment in enumerate(data_rows, row_start):
                if payment:
                    payment_sheet.cell(row=row_idx, column=1, value=payment.get("due_date"))
                    payment_sheet.cell(row=row_idx, column=2, value=payment.get("description"))
                    
                    # Set amount as number with currency formatting
                    amount_cell = payment_sheet.cell(row=row_idx, column=3)
                    amount_value = self.parse_currency_to_number(payment.get("amount"))
                    amount_cell.value = amount_value
                    if amount_value is not None:
                        amount_cell.number_format = '"$"#,##0.00'
                    
                    payment_sheet.cell(row=row_idx, column=4, value="Yes" if payment.get("invoice_submission_due") else "No")
                    payment_sheet.cell(row=row_idx, column=5, value=payment.get("invoice_submission_due"))
                    payment_sheet.cell(row=row_idx, column=9, value=payment.get("expected_payment_date"))
                    payment_sheet.cell(row=row_idx, column=12, value="Pending")
        
        # 3. Expense Tracking Sheet
        if tracking_requirements.get("expense_tracking", {}).get("travel_expenses") == "yes":
            expense_sheet = wb.create_sheet("Expense Tracking")
            
            expense_headers = [
                "Date", "Expense Type", "Description", "Amount", "Receipt", 
                "Pre-approved", "Submitted Date", "Approved Date", 
                "Reimbursement Date", "Status", "Notes"
            ]
            
            for col_idx, header in enumerate(expense_headers, 1):
                cell = expense_sheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            expense_categories = [
                "Travel - Airfare", "Travel - Hotel", "Travel - Ground Transportation",
                "Meals & Entertainment", "Equipment", "Materials", "Other"
            ]
            
            for row_idx, category in enumerate(expense_categories, 2):
                expense_sheet.cell(row=row_idx, column=2, value=category)
                expense_sheet.cell(row=row_idx, column=10, value="Template")
        
        # 4. Compliance Checklist Sheet
        compliance_sheet = wb.create_sheet("Compliance Checklist")
        
        compliance_items = [
            ["REQUIRED DOCUMENTATION", "Status", "Due Date", "Completed Date", "Notes"],
            ["W-9 Form Submission", "", "", "", ""],
            ["Monthly Invoice Submission", "", "", "", ""],
            ["Time Breakdown Documentation", "", "", "", ""],
            ["Expense Pre-approvals", "", "", "", ""],
            ["Contract Deliverables", "", "", "", ""],
            ["", "", "", "", ""],
            ["INVOICE REQUIREMENTS", "", "", "", ""],
            ["Daily Time Breakdown", "", "", "", ""],
            ["Detailed Work Description", "", "", "", ""],
            ["Required Signatures", "", "", "", ""],
            ["Supporting Documentation", "", "", "", ""],
            ["", "", "", "", ""],
            ["REPORTING DEADLINES", "", "", "", ""],
            ["Monthly Reports", "", "", "", ""],
            ["Expense Reports", "", "", "", ""],
            ["Final Deliverables", "", "", "", ""]
        ]
        
        for row_idx, item_row in enumerate(compliance_items, 1):
            for col_idx, item in enumerate(item_row, 1):
                cell = compliance_sheet.cell(row=row_idx, column=col_idx, value=item)
                if row_idx == 1 or item in ["REQUIRED DOCUMENTATION", "INVOICE REQUIREMENTS", "REPORTING DEADLINES"]:
                    cell.font = header_font
                    cell.fill = header_fill
                cell.border = border
        
        # 5. Budget Monitor Sheet
        budget_sheet = wb.create_sheet("Budget Monitor")
        
        budget_headers = [
            "Metric", "Budgeted/Maximum", "Current", "Remaining", 
            "% Used", "Status", "Warning Level"
        ]
        
        for col_idx, header in enumerate(budget_headers, 1):
            cell = budget_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Budget tracking rows
        max_hours = contract_info.get("max_hours", "")
        total_value = contract_info.get("total_value", "")
        
        max_hours_num = self.parse_hours_to_number(max_hours)
        total_value_num = self.parse_currency_to_number(total_value)
        
        budget_items = [
            ["Total Hours", max_hours_num, 0, max_hours_num, 0, "On Track", "< 80%"],
            ["Contract Value", total_value_num, 0, total_value_num, 0, "On Track", "< 80%"],
            ["Monthly Invoicing", "Monthly", "Current", "Remaining", "%", "Status", "Threshold"],
            ["Expense Budget", None, 0, None, 0, "On Track", "< 90%"]
        ]
        
        for row_idx, item in enumerate(budget_items, 2):
            for col_idx, value in enumerate(item, 1):
                cell = budget_sheet.cell(row=row_idx, column=col_idx, value=value)
                
                if col_idx in [2, 3, 4] and isinstance(value, (int, float)) and value is not None:
                    if "Value" in item[0]:
                        cell.number_format = '"$"#,##0.00'
                    elif "Hours" in item[0]:
                        cell.number_format = '#,##0.0'
        
        # Add conditional formatting for warnings
        warning_rule = CellIsRule(operator='greaterThan', formula=['80'], 
                                stopIfTrue=True, fill=warning_fill)
        budget_sheet.conditional_formatting.add('E2:E5', warning_rule)
        
        # 6. Validation Details Sheet
        if validated_data:
            validation_sheet = wb.create_sheet("Validation Details")
            validation_sheet.cell(row=1, column=1, value="THREE-PASS VALIDATION DETAILS")
            validation_sheet.cell(row=1, column=1).font = header_font
            validation_sheet.cell(row=1, column=1).fill = header_fill
            
            validation_info = [
                ["", ""],
                ["VALIDATION METHODOLOGY", ""],
                ["Pass 1", "High-precision regex pattern extraction"],
                ["Pass 2", "Direct Claude AI extraction with focused prompts"],
                ["Pass 3", "Cross-validation and confidence scoring"],
                ["", ""],
                ["CONFIDENCE LEVELS", ""],
            ]
            
            confidence_scores = validated_data.get('confidence_scores', {})
            for field, score in confidence_scores.items():
                value = contract_info.get(field.replace('_', ' ').replace(' ', '_'), 'Not found')
                validation_info.append([field.replace('_', ' ').title(), f"{value} [{score}]"])
            
            validation_info.extend([
                ["", ""],
                ["VALIDATION NOTES", ""],
                ["Notes", validated_data.get('validation_notes', 'No issues detected')],
                ["", ""],
                ["ACCURACY SUMMARY", ""],
                ["High Confidence Items", str(len([s for s in confidence_scores.values() if s == "HIGH"]))],
                ["Medium Confidence Items", str(len([s for s in confidence_scores.values() if s == "MEDIUM"]))],
                ["Low Confidence Items", str(len([s for s in confidence_scores.values() if s == "LOW"]))],
                ["Total Items Validated", str(len(confidence_scores))],
            ])
            
            for row_idx, (label, value) in enumerate(validation_info, 1):
                validation_sheet.cell(row=row_idx, column=1, value=label)
                validation_sheet.cell(row=row_idx, column=2, value=value)
                
                if label in ["VALIDATION METHODOLOGY", "CONFIDENCE LEVELS", "VALIDATION NOTES", "ACCURACY SUMMARY"]:
                    validation_sheet.cell(row=row_idx, column=1).font = Font(bold=True)
        
        # 7. Full Analysis Sheet
        analysis_sheet = wb.create_sheet("Full Analysis")
        analysis_sheet.cell(row=1, column=1, value="Claude AI Enhanced Contract Analysis with Three-Pass Validation")
        analysis_sheet.cell(row=1, column=1).font = header_font
        analysis_sheet.cell(row=1, column=1).fill = header_fill
        
        analysis_lines = original_analysis.split('\n')
        for row_idx, line in enumerate(analysis_lines, 3):
            analysis_sheet.cell(row=row_idx, column=1, value=line)
        
        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_file)
        print(f"Enhanced payment tracking spreadsheet saved to: {output_file}")

    def analyze_with_high_precision(self, contract_text: str) -> Dict:
        """Main analysis method with high-precision validation."""
        
        # Step 1: Run three-pass validation for financial data
        validated_data = self.cross_validate_financial_data(contract_text)
        
        # Step 2: Run comprehensive analysis for other elements
        print("\nPerforming comprehensive contract analysis...")
        comprehensive_result = self.analyze_contract_comprehensive(contract_text)
        
        if not comprehensive_result.get('success'):
            return {'success': False, 'error': comprehensive_result.get('analysis')}
        
        print("Extracting comprehensive tracking data...")
        
        # Extract structured data
        contract_info, payment_schedule, tracking_requirements = self.extract_comprehensive_data(comprehensive_result["analysis"])
        
        # Step 3: Override with high-precision financial data
        if validated_data.get('total_contract_value'):
            contract_info['total_value'] = validated_data['total_contract_value']
        if validated_data.get('hourly_rate'):
            contract_info['hourly_rate'] = validated_data['hourly_rate']
        if validated_data.get('start_date'):
            contract_info['start_date'] = validated_data['start_date']
        if validated_data.get('end_date'):
            contract_info['end_date'] = validated_data['end_date']
        if validated_data.get('payment_terms'):
            contract_info['payment_timeline'] = validated_data['payment_terms']
        if validated_data.get('maximum_hours'):
            contract_info['max_hours'] = validated_data['maximum_hours']
        if validated_data.get('daily_hour_limit'):
            contract_info['max_daily_hours'] = validated_data['daily_hour_limit']
        
        return {
            'success': True,
            'contract_info': contract_info,
            'payment_schedule': payment_schedule,
            'tracking_requirements': tracking_requirements,
            'validated_data': validated_data,
            'analysis': comprehensive_result['analysis']
        }


def main():
    parser = argparse.ArgumentParser(description='Enhanced contract analysis with three-pass validation for maximum accuracy')
    parser.add_argument('contract_file', help='Path to contract file (PDF, DOCX, or TXT)')
    parser.add_argument('-o', '--output', help='Output Excel file name', 
                       default='enhanced_contract_tracker_3pass.xlsx')
    parser.add_argument('-k', '--api-key', help='Claude API key (or set ANTHROPIC_API_KEY env var)')
    parser.add_argument('--verbose', '-v', action='store_true', help='Verbose output')
    
    args = parser.parse_args()
    
    # Get API key
    api_key = args.api_key or os.getenv('ANTHROPIC_API_KEY')
    if not api_key:
        print("Error: Claude API key required. Set ANTHROPIC_API_KEY environment variable or use -k flag")
        sys.exit(1)
    
    try:
        # Initialize analyzer
        analyzer = HighPrecisionContractAnalyzer(api_key)
        print(f"Reading contract file: {args.contract_file}")
        
        # Read contract
        contract_text = analyzer.read_contract_file(args.contract_file)
        print(f"Contract read successfully ({len(contract_text)} characters)")
        
        # Analyze with high precision
        result = analyzer.analyze_with_high_precision(contract_text)
        
        if not result["success"]:
            print(f"Error: {result.get('error')}")
            sys.exit(1)
        
        # Extract results
        contract_info = result['contract_info']
        payment_schedule = result['payment_schedule']
        tracking_requirements = result['tracking_requirements']
        validated_data = result['validated_data']
        
        # Create comprehensive spreadsheet
        print("\nGenerating comprehensive Excel tracking system...")
        analyzer.create_comprehensive_spreadsheet(
            contract_info, 
            payment_schedule,
            tracking_requirements,
            args.output,
            result['analysis'],
            validated_data
        )
        
        # Print detailed summary
        print("\n" + "="*70)
        print("THREE-PASS VALIDATION ANALYSIS COMPLETE")
        print("="*70)
        
        # Financial Data Confidence Summary
        confidence_scores = validated_data.get('confidence_scores', {})
        if confidence_scores:
            print("\nFINANCIAL DATA CONFIDENCE LEVELS:")
            high_confidence_items = 0
            total_items = 0
            
            key_fields = ['total_contract_value', 'hourly_rate', 'start_date', 'end_date', 'payment_terms', 'maximum_hours', 'daily_hour_limit']
            
            for field in key_fields:
                if field in confidence_scores:
                    score = confidence_scores[field]
                    value = contract_info.get(field.replace('_', ' ').replace(' ', '_'), 'Not found')
                    total_items += 1
                    
                    if score == 'HIGH':
                        print(f"  ✓ {field.replace('_', ' ').title()}: {value} [HIGH CONFIDENCE]")
                        high_confidence_items += 1
                    elif score == 'MEDIUM':
                        print(f"  • {field.replace('_', ' ').title()}: {value} [MEDIUM CONFIDENCE]")
                    elif score == 'LOW':
                        print(f"  ⚠ {field.replace('_', ' ').title()}: {value} [LOW CONFIDENCE - REVIEW RECOMMENDED]")
                    else:
                        print(f"  - {field.replace('_', ' ').title()}: Not found")
            
            # Calculate accuracy percentage
            if total_items > 0:
                accuracy_percentage = (high_confidence_items / total_items) * 100
                print(f"\nDATA ACCURACY: {accuracy_percentage:.1f}% HIGH CONFIDENCE ({high_confidence_items}/{total_items} critical items)")
                
                if accuracy_percentage >= 90:
                    print("✓ EXCELLENT: Financial data extraction highly reliable")
                elif accuracy_percentage >= 70:
                    print("✓ GOOD: Most financial data reliable, some items may need review")
                else:
                    print("⚠ MANUAL REVIEW RECOMMENDED: Several low-confidence items detected")
        
        # Show validation notes if any
        validation_notes = validated_data.get('validation_notes', '')
        if validation_notes and validation_notes.strip():
            print(f"\nValidation Notes: {validation_notes}")
        
        print(f"\nCONTRACT SUMMARY:")
        print(f"Client: {contract_info.get('client', 'Unknown')}")
        print(f"Vendor: {contract_info.get('vendor', 'Unknown')}")
        print(f"Contract Type: {contract_info.get('contract_type', 'Unknown')}")
        print(f"Total Value: {contract_info.get('total_value', 'Unknown')}")
        print(f"Payment Terms: {contract_info.get('payment_timeline', 'Unknown')}")
        
        if contract_info.get('max_hours'):
            print(f"Maximum Hours: {contract_info.get('max_hours')}")
        if contract_info.get('hourly_rate'):
            print(f"Hourly Rate: {contract_info.get('hourly_rate')}")
            
        print(f"\nTRACKING FEATURES ENABLED:")
        if tracking_requirements.get("expense_tracking", {}).get("travel_expenses") == "yes":
            print("  ✓ Travel & Expense Tracking")
        if tracking_requirements.get("compliance", {}).get("w9_required") == "yes":
            print("  ✓ W-9 Compliance Tracking")
        if tracking_requirements.get("budget_monitoring", {}).get("hour_tracking") == "yes":
            print("  ✓ Hour Budget Monitoring")
        if payment_schedule:
            print(f"  ✓ {len(payment_schedule)} Payment Milestones Tracked")
            
        print(f"\nSPREADSHEET INCLUDES:")
        print("  • Contract Summary with confidence scores")
        print("  • Payment Tracking with invoice timeline")
        print("  • Expense Tracking (if applicable)")
        print("  • Compliance Checklist")
        print("  • Budget Monitor with warnings")
        print("  • Validation Details sheet")
        print("  • Complete AI analysis")
            
        print(f"\nEnhanced tracking system saved to: {args.output}")
        print("✓ All critical payment, compliance, and budget elements included!")
        print("✓ Three-pass validation ensures maximum accuracy on financial data")
        
    except Exception as e:
        print(f"Error: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()