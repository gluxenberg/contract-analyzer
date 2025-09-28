#!/usr/bin/env python3
"""
Contract Financial Analyzer - NEW VERSION (with working timer & progress)
Analysis-driven approach with NO hardcoded assumptions.
Payment schedules generated ONLY from actual contract terms.

Changes in this version:
- Determinate progress bar with clear phase percentages
- Elapsed timer label (mm:ss) that runs during long operations
- Thread-safe UI updates via root.after()
- Non-blocking pulse during API call to show forward movement up to a ceiling
- Progress + timer also used during Export
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
import json
import re
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import anthropic
from typing import Dict, List, Optional
import PyPDF2
import docx


class ContractAnalyzer:
    def __init__(self, api_key: str):
        """Initialize contract analyzer with Claude API - NO hardcoded assumptions."""
        self.client = anthropic.Anthropic(api_key=api_key)

    def read_contract_file(self, file_path: str) -> str:
        """Read contract file and return content as string."""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Contract file not found: {file_path}")

        try:
            if path.suffix.lower() == '.pdf':
                with open(file_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    text = ""
                    for page in pdf_reader.pages:
                        t = page.extract_text() or ""
                        text += t + "\n"
                    return text

            elif path.suffix.lower() in ['.docx', '.doc']:
                doc = docx.Document(file_path)
                text = ""
                for paragraph in doc.paragraphs:
                    text += paragraph.text + "\n"
                return text

            elif path.suffix.lower() == '.txt':
                with open(file_path, 'r', encoding='utf-8') as file:
                    return file.read()

            else:
                raise ValueError(f"Unsupported file format: {path.suffix}")

        except Exception as e:
            raise Exception(f"Error reading contract file: {str(e)}")

    def analyze_contract_comprehensive(self, contract_text: str) -> Dict:
        """
        Comprehensive contract analysis using Claude API.
        NO assumptions - only extract what's explicitly stated.
        """

        analysis_prompt = f"""
        Please analyze this contract and extract ONLY the information that is explicitly stated.
        Do not make assumptions or infer information that is not directly written in the contract.

        IMPORTANT: If payment terms, schedules, or amounts are not explicitly specified,
        report "Not specified in contract" - do NOT create payment schedules based on assumptions.

        Extract the following information:

        1. BASIC CONTRACT INFORMATION:
           - Contract title/type
           - Parties involved (contractor and client)
           - Contract start and end dates
           - Total contract value (if explicitly stated)

        2. PAYMENT TERMS (extract ONLY what is explicitly stated):
           - Payment schedule (monthly, quarterly, milestone-based, etc.)
           - Specific payment amounts and due dates
           - Payment conditions or triggers
           - Late payment penalties
           - Payment methods accepted

        3. DELIVERABLES AND MILESTONES:
           - Specific deliverables mentioned
           - Delivery dates
           - Milestone payments tied to deliverables
           - Performance criteria

        4. SPECIAL FINANCIAL TERMS:
           - Deposits, retainers, or upfront payments
           - Hourly rates (if time & materials)
           - Expense reimbursement terms
           - Cancellation fees or penalties

        5. OTHER IMPORTANT TERMS:
           - Termination clauses
           - Dispute resolution
           - Governing law
           - Any special conditions

        Return as JSON with this structure:
        {{
            "contract_info": {{
                "title": "contract title or null",
                "contractor": "contractor name or null",
                "client": "client name or null",
                "start_date": "YYYY-MM-DD or null",
                "end_date": "YYYY-MM-DD or null",
                "total_value": "exact amount with currency or null"
            }},
            "payment_terms": {{
                "schedule_type": "monthly/quarterly/milestone/lump_sum/not_specified",
                "payment_frequency": "description of frequency or null",
                "specific_payments": [
                    {{
                        "amount": "exact amount",
                        "due_date": "date or condition",
                        "description": "what triggers this payment"
                    }}
                ],
                "payment_conditions": "any conditions for payment or null",
                "late_payment_penalty": "penalty terms or null",
                "payment_methods": "accepted payment methods or null"
            }},
            "deliverables": [
                {{
                    "description": "deliverable description",
                    "due_date": "delivery date or null",
                    "associated_payment": "payment amount if tied to this deliverable"
                }}
            ],
            "financial_details": {{
                "deposits": "deposit amounts or null",
                "hourly_rate": "rate if T&M or null",
                "expense_reimbursement": "expense terms or null",
                "cancellation_fees": "cancellation terms or null"
            }},
            "contract_analysis": {{
                "payment_schedule_defined": true/false,
                "deliverables_specified": true/false,
                "completion_timeline_clear": true/false,
                "payment_triggers_clear": true/false
            }}
        }}

        Contract text:
        {contract_text}
        """

        try:
            response = self.client.messages.create(
                model="claude-3-5-haiku-20241022",
                max_tokens=2000,
                messages=[{"role": "user", "content": analysis_prompt}]
            )

            response_text = response.content[0].text

            # Extract JSON from response
            start_idx = response_text.find('{')
            end_idx = response_text.rfind('}') + 1

            if start_idx != -1 and end_idx > start_idx:
                analysis_result = json.loads(response_text[start_idx:end_idx])
                return analysis_result
            else:
                return {"error": "Could not parse analysis response"}

        except Exception as e:
            return {"error": f"Analysis failed: {str(e)}"}

    def generate_payment_schedule_from_analysis(self, analysis: Dict) -> List[Dict]:
        """
        Generate payment schedule ONLY if explicitly defined in contract analysis.
        NO assumptions or hardcoded patterns.
        """

        if "error" in analysis:
            return []

        payment_terms = analysis.get("payment_terms", {})
        specific_payments = payment_terms.get("specific_payments", [])

        # If contract has specific payment schedule, use it
        if specific_payments:
            schedule = []
            for payment in specific_payments:
                schedule.append({
                    "amount": payment.get("amount", "Amount not specified"),
                    "due_date": payment.get("due_date", "Date not specified"),
                    "description": payment.get("description", "Payment"),
                    "source": "Contract specification"
                })
            return schedule

        # If no specific payments but payment schedule type is defined
        schedule_type = payment_terms.get("schedule_type", "not_specified")
        total_value = analysis.get("contract_info", {}).get("total_value")

        if schedule_type == "not_specified" or not total_value:
            return []  # Cannot create schedule without explicit terms

        # Only create schedule if contract explicitly defines the pattern
        if schedule_type == "lump_sum":
            return [{
                "amount": total_value,
                "due_date": "As specified in contract",
                "description": "Lump sum payment",
                "source": "Contract specification"
            }]

        # For other types, we need explicit payment details from the contract
        # Do not assume payment amounts or dates
        return []

    def _parse_currency_to_number(self, currency_string):
        """Convert currency string to number for Excel."""
        if not currency_string or currency_string in ['null', 'N/A', 'Not found', '', 'Not specified']:
            return None

        s = str(currency_string).strip()

        # Handle negatives in parentheses e.g. "(1,234.56)"
        paren = re.search(r'\(\s*\$?\s*(-?\d{1,3}(?:,\d{3})*(?:\.\d+)?|-?\d+(?:\.\d+)?)\s*\)', s)
        if paren:
            try:
                return -float(paren.group(1).replace(',', ''))
            except ValueError:
                pass

        # 1) Prefer a $-prefixed currency amount
        m = re.search(r'\$\s*(-?\d{1,3}(?:,\d{3})*(?:\.\d+)?|-?\d+(?:\.\d+)?)', s)
        if not m:
            # 2) Otherwise take the first number not followed by %
            m = re.search(r'(-?\d{1,3}(?:,\d{3})*(?:\.\d+)?|-?\d+(?:\.\d+)?)(?!\s*%)', s)

        if not m:
            return None

        value_str = m.group(1).replace(',', '')

        try:
            return float(value_str)
        except (ValueError, TypeError):
            return None

    def _format_excel_value(self, key, value):
        """Format value appropriately for Excel - numbers as numbers, text as text."""
        if not value or value in ['null', 'N/A', 'Not found', '']:
            return "Not specified"

        # Check if this should be a number based on key name
        number_keys = ['value', 'amount', 'rate', 'total', 'deposit', 'retainer', 'fee', 'cost', 'price', 'budget']

        if any(num_key in key.lower() for num_key in number_keys):
            # Try to parse as number
            parsed_number = self._parse_currency_to_number(value)
            if parsed_number is not None:
                return parsed_number

        return str(value)

    def export_analysis_to_excel(self, analysis: Dict, payment_schedule: List[Dict], file_path: str):
        """Export contract analysis and payment schedule to Excel with proper number formatting."""

        wb = Workbook()

        # Contract Analysis Sheet
        ws1 = wb.active
        ws1.title = "Contract Analysis"

        # Headers
        header_font = Font(bold=True, size=12)
        ws1['A1'] = "Contract Analysis Report"
        ws1['A1'].font = Font(bold=True, size=14)

        row = 3

        # Contract Information
        if "contract_info" in analysis:
            ws1[f'A{row}'] = "CONTRACT INFORMATION"
            ws1[f'A{row}'].font = header_font
            row += 1

            contract_info = analysis["contract_info"]
            for key, value in contract_info.items():
                ws1[f'A{row}'] = key.replace('_', ' ').title()
                formatted_value = self._format_excel_value(key, value)
                cell = ws1[f'B{row}']
                cell.value = formatted_value

                # Format as number if it's numeric
                if isinstance(formatted_value, (int, float)):
                    cell.number_format = '#,##0.00'

                row += 1
            row += 1

        # Payment Terms
        if "payment_terms" in analysis:
            ws1[f'A{row}'] = "PAYMENT TERMS"
            ws1[f'A{row}'].font = header_font
            row += 1

            payment_terms = analysis["payment_terms"]
            for key, value in payment_terms.items():
                if key != "specific_payments":  # Handle this separately
                    ws1[f'A{row}'] = key.replace('_', ' ').title()
                    formatted_value = self._format_excel_value(key, value)
                    cell = ws1[f'B{row}']
                    cell.value = formatted_value

                    # Format as number if it's numeric
                    if isinstance(formatted_value, (int, float)):
                        cell.number_format = '#,##0.00'

                    row += 1
            row += 1

        # Financial Details
        if "financial_details" in analysis:
            ws1[f'A{row}'] = "FINANCIAL DETAILS"
            ws1[f'A{row}'].font = header_font
            row += 1

            financial_details = analysis["financial_details"]
            for key, value in financial_details.items():
                ws1[f'A{row}'] = key.replace('_', ' ').title()
                formatted_value = self._format_excel_value(key, value)
                cell = ws1[f'B{row}']
                cell.value = formatted_value

                # Format as number if it's numeric
                if isinstance(formatted_value, (int, float)):
                    cell.number_format = '#,##0.00'

                row += 1

        # Payment Schedule Sheet
        if payment_schedule:
            ws2 = wb.create_sheet("Payment Schedule")

            # Headers - new order: Description, Due Date, Source, Amount
            headers = ["Description", "Due Date", "Source", "Amount"]
            for col, header in enumerate(headers, 1):
                ws2.cell(1, col, header).font = header_font

            # Payment data with proper number formatting and new column order
            for row_idx, payment in enumerate(payment_schedule, 2):
                # Column 1: Description
                ws2.cell(row_idx, 1, payment.get("description", ""))

                # Column 2: Due Date
                ws2.cell(row_idx, 2, payment.get("due_date", ""))

                # Column 3: Source
                ws2.cell(row_idx, 3, payment.get("source", ""))

                # Column 4: Amount - convert to number
                amount_str = payment.get("amount", "")
                amount_num = self._parse_currency_to_number(amount_str)
                amount_cell = ws2.cell(row_idx, 4, amount_num if amount_num is not None else amount_str)

                # Format amount cell as currency if it's a number
                if amount_num is not None:
                    amount_cell.number_format = '#,##0.00'  # Number with thousands separator

            # No total row - removed per requirements
        else:
            ws2 = wb.create_sheet("Payment Schedule")
            ws2['A1'] = "No payment schedule specified in contract"
            ws2['A1'].font = header_font

        # Auto-adjust column widths
        for ws in wb.worksheets:
            max_col = ws.max_column or 1
            for col_idx in range(1, max_col + 1):
                max_len = 0
                for col in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
                    for cell in col:
                        try:
                            if cell.value is not None:
                                val_len = len(str(cell.value))
                                if val_len > max_len:
                                    max_len = val_len
                        except Exception:
                            pass
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        wb.save(file_path)


class ContractAnalyzerGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Contract Analyzer - Analysis-Driven (No Assumptions)")
        self.root.geometry("1000x740")

        self.analyzer = None
        self.current_analysis = None
        self.current_schedule = None

        # API key storage
        self.api_key = tk.StringVar()

        # Progress tracking
        self.start_time: Optional[datetime] = None
        self.timer_id: Optional[str] = None
        self.pulse_id: Optional[str] = None
        self.progress_ceiling = 0  # max percent the pulse may reach during a phase

        self.setup_gui()
        self.load_saved_api_key()

    # ---------------- GUI -----------------
    def setup_gui(self):
        """Setup the GUI interface."""

        # Main frame
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # API Key section
        api_frame = ttk.LabelFrame(main_frame, text="Claude API Configuration", padding=10)
        api_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(api_frame, text="API Key:").pack(side=tk.LEFT)
        self.api_key_entry = ttk.Entry(api_frame, textvariable=self.api_key, show="*", width=50)
        self.api_key_entry.pack(side=tk.LEFT, padx=(5, 10))

        ttk.Button(api_frame, text="Save", command=self.save_api_key).pack(side=tk.LEFT, padx=(5, 0))
        ttk.Button(api_frame, text="Set API Key", command=self.set_api_key).pack(side=tk.LEFT, padx=(5, 0))
        ttk.Button(api_frame, text="Test Connection", command=self.test_api_connection).pack(side=tk.LEFT)

        # File selection
        file_frame = ttk.LabelFrame(main_frame, text="Contract Analysis", padding=10)
        file_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Button(file_frame, text="Select Contract File", command=self.select_file).pack(side=tk.LEFT)
        self.file_label = ttk.Label(file_frame, text="No file selected")
        self.file_label.pack(side=tk.LEFT, padx=(10, 0))

        # Analysis button
        self.analyze_button = ttk.Button(file_frame, text="Analyze Contract",
                                         command=self.analyze_contract, state=tk.DISABLED)
        self.analyze_button.pack(side=tk.RIGHT)

        # Results area
        results_frame = ttk.LabelFrame(main_frame, text="Analysis Results", padding=10)
        results_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # Create notebook for tabs
        self.notebook = ttk.Notebook(results_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # Contract Analysis tab
        self.analysis_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.analysis_frame, text="Contract Analysis")

        self.analysis_text = scrolledtext.ScrolledText(self.analysis_frame, wrap=tk.WORD)
        self.analysis_text.pack(fill=tk.BOTH, expand=True)

        # Payment Schedule tab
        self.schedule_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.schedule_frame, text="Payment Schedule")

        self.schedule_text = scrolledtext.ScrolledText(self.schedule_frame, wrap=tk.WORD)
        self.schedule_text.pack(fill=tk.BOTH, expand=True)

        # Export section
        export_frame = ttk.LabelFrame(main_frame, text="Export Results", padding=10)
        export_frame.pack(fill=tk.X)

        self.export_button = ttk.Button(export_frame, text="Export to Excel",
                                        command=self.export_results, state=tk.DISABLED)
        self.export_button.pack(side=tk.LEFT)

        # PROGRESS UI
        prog_frame = ttk.Frame(main_frame)
        prog_frame.pack(fill=tk.X, pady=(12, 4))

        ttk.Label(prog_frame, text="Progress:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky='w')
        self.progress_bar = ttk.Progressbar(prog_frame, length=600, mode='determinate', maximum=100)
        self.progress_bar.grid(row=0, column=1, sticky='we', padx=(8, 0))
        prog_frame.columnconfigure(1, weight=1)

        # Elapsed timer label
        self.elapsed_var = tk.StringVar(value="00:00")
        ttk.Label(prog_frame, textvariable=self.elapsed_var, width=6, anchor='e').grid(row=0, column=2, padx=(8, 0))

        # Status bar
        self.status_var = tk.StringVar()
        self.status_var.set("Ready - No assumptions, analysis-driven only")
        tk.Label(main_frame, textvariable=self.status_var).pack(fill=tk.X, pady=(5, 0))

    # ------------- API Key Mgmt -------------
    def load_saved_api_key(self):
        """Load saved API key from config file."""
        try:
            config_dir = Path.home() / ".contract_analyzer"
            config_file = config_dir / "config.txt"

            if config_file.exists():
                with open(config_file, 'r') as f:
                    saved_key = f.read().strip()
                    if saved_key:
                        self.api_key.set(saved_key)
                        # Auto-initialize analyzer if key exists
                        try:
                            self.analyzer = ContractAnalyzer(saved_key)
                            self.analyze_button.config(state=tk.NORMAL if hasattr(self, 'selected_file') else tk.DISABLED)
                        except:
                            pass  # Key might be invalid
        except Exception:
            pass  # Config file might not exist or be readable

    def save_api_key(self):
        """Save API key to config file."""
        try:
            api_key = self.api_key.get().strip()
            if not api_key:
                messagebox.showwarning("Warning", "Please enter an API key first")
                return

            config_dir = Path.home() / ".contract_analyzer"
            config_dir.mkdir(exist_ok=True)

            config_file = config_dir / "config.txt"
            with open(config_file, 'w') as f:
                f.write(api_key)

            messagebox.showinfo("Success", "API key saved successfully")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save API key: {str(e)}")

    def set_api_key(self):
        """Set the Claude API key."""
        api_key = self.api_key.get().strip()
        if not api_key:
            messagebox.showerror("Error", "Please enter a valid API key")
            return

        # Basic API key format validation
        if not api_key.startswith('sk-ant-'):
            messagebox.showerror("Error", "API key should start with 'sk-ant-'")
            return

        if len(api_key) < 50:
            messagebox.showerror("Error", "API key appears to be too short. Please check the full key.")
            return

        try:
            self.analyzer = ContractAnalyzer(api_key)
            messagebox.showinfo("Success", "API key set successfully")
            self.analyze_button.config(state=tk.NORMAL if hasattr(self, 'selected_file') else tk.DISABLED)
        except Exception as e:
            error_msg = str(e)
            if "authentication_error" in error_msg:
                messagebox.showerror("Authentication Error",
                    "Invalid API key. Please check:\n"
                    "• Key starts with 'sk-ant-'\n"
                    "• Key is complete (not truncated)\n"
                    "• Key is active and not expired\n"
                    "• Key has proper permissions")
            else:
                messagebox.showerror("Error", f"Failed to initialize analyzer: {error_msg}")

    def test_api_connection(self):
        """Test API connection with a simple request."""
        api_key = self.api_key.get().strip()
        if not api_key:
            messagebox.showerror("Error", "Please enter an API key first")
            return

        if not api_key.startswith('sk-ant-'):
            messagebox.showerror("Error", "API key should start with 'sk-ant-'")
            return

        try:
            # Create a test analyzer
            test_analyzer = ContractAnalyzer(api_key)

            # Make a simple test request
            _ = test_analyzer.client.messages.create(
                model="claude-3-5-haiku-20241022",
                max_tokens=10,
                messages=[{"role": "user", "content": "Test"}]
            )

            messagebox.showinfo("Success",
                "API connection successful!\n"
                "Your API key is valid and working.")

        except Exception as e:
            error_msg = str(e)
            if "authentication_error" in error_msg:
                messagebox.showerror("Authentication Error",
                    "API key is invalid. Please check:\n"
                    "• Key is copied correctly from Anthropic Console\n"
                    "• Key starts with 'sk-ant-'\n"
                    "• Key has not expired\n"
                    "• Account has sufficient credits")
            elif "permission" in error_msg.lower():
                messagebox.showerror("Permission Error",
                    "API key lacks required permissions.\n"
                    "Please check your Anthropic account settings.")
            else:
                messagebox.showerror("Connection Error", f"Failed to connect: {error_msg}")

    # ------------- File & Analysis -------------
    def select_file(self):
        """Select contract file for analysis."""
        file_path = filedialog.askopenfilename(
            title="Select Contract File",
            filetypes=[
                ("All Supported", "*.pdf *.docx *.doc *.txt"),
                ("PDF files", "*.pdf"),
                ("Word files", "*.docx *.doc"),
                ("Text files", "*.txt")
            ]
        )

        if file_path:
            self.selected_file = file_path
            self.file_label.config(text=os.path.basename(file_path))
            self.analyze_button.config(state=tk.NORMAL if self.analyzer else tk.DISABLED)

            # Clear previous results when new file is selected
            self.clear_results()
            self.status_var.set(f"New file selected: {os.path.basename(file_path)}")

    def analyze_contract(self):
        """Analyze the selected contract."""
        if not self.analyzer or not hasattr(self, 'selected_file'):
            messagebox.showerror("Error", "Please set API key and select a file first")
            return

        # Disable analyze button during analysis
        self.analyze_button.config(state=tk.DISABLED)
        self.export_button.config(state=tk.DISABLED)

        # Run analysis in separate thread to prevent GUI freezing
        thread = threading.Thread(target=self._run_analysis)
        thread.daemon = True
        thread.start()

    # Progress helpers
    def _start_timer(self):
        self.start_time = datetime.now()
        self._tick_timer()

    def _tick_timer(self):
        if self.start_time is None:
            return
        elapsed = int((datetime.now() - self.start_time).total_seconds())
        mins, secs = divmod(elapsed, 60)
        self.elapsed_var.set(f"{mins:02d}:{secs:02d}")
        self.timer_id = self.root.after(1000, self._tick_timer)

    def _stop_timer(self):
        if self.timer_id:
            self.root.after_cancel(self.timer_id)
            self.timer_id = None
        self.start_time = None

    def _set_status(self, msg: str):
        self.status_var.set(msg)

    def _set_progress(self, pct: int):
        pct = max(0, min(100, pct))
        self.progress_bar['value'] = pct

    def _start_pulse_towards(self, ceiling: int, step: int = 1, interval_ms: int = 250):
        """During long phases (e.g., API call), gently advance progress up to ceiling."""
        self.progress_ceiling = max(0, min(100, ceiling))

        def _pulse():
            current = int(self.progress_bar['value'])
            if current < self.progress_ceiling:
                self.progress_bar['value'] = min(self.progress_ceiling, current + step)
                self.pulse_id = self.root.after(interval_ms, _pulse)
            else:
                self.pulse_id = None
        # start pulse
        if self.pulse_id:
            self.root.after_cancel(self.pulse_id)
        _pulse()

    def _stop_pulse(self):
        if self.pulse_id:
            self.root.after_cancel(self.pulse_id)
            self.pulse_id = None

    def _run_analysis(self):
        """Run contract analysis in background thread."""
        try:
            # Initialize progress & timer
            self.root.after(0, lambda: (self._set_progress(0), self._set_status("Starting analysis...")))
            self.root.after(0, self._start_timer)

            # Phase 1: Read file (0 -> 20%)
            self.root.after(0, lambda: self._set_status("Reading contract file..."))
            contract_text = self.analyzer.read_contract_file(self.selected_file)
            self.root.after(0, lambda: self._set_progress(20))

            # Phase 2: API analysis (20% -> soft 85% with pulse)
            self.root.after(0, lambda: self._set_status("Analyzing contract with Claude API..."))
            self.root.after(0, lambda: self._start_pulse_towards(85, step=1, interval_ms=300))
            analysis = self.analyzer.analyze_contract_comprehensive(contract_text)
            self.root.after(0, self._stop_pulse)
            self.root.after(0, lambda: self._set_progress(85))

            # Phase 3: Generate schedule (85% -> 92%)
            self.root.after(0, lambda: self._set_status("Generating payment schedule from analysis..."))
            payment_schedule = self.analyzer.generate_payment_schedule_from_analysis(analysis)
            self.root.after(0, lambda: self._set_progress(92))

            # Finish (-> 100%)
            def finish_ui():
                self._display_results(analysis, payment_schedule)
                self._set_progress(100)
                self._set_status("Analysis complete - No assumptions made")
                self._stop_timer()
            self.root.after(0, finish_ui)

        except Exception as e:
            def on_err():
                self._stop_pulse()
                self._stop_timer()
                error_msg = f"Analysis failed: {str(e)}"
                messagebox.showerror("Error", error_msg)
                self._set_status("Analysis failed")
                self.analyze_button.config(state=tk.NORMAL)
            self.root.after(0, on_err)

    def clear_results(self):
        """Clear all analysis results from the display."""
        try:
            # Clear text areas completely
            self.analysis_text.delete(1.0, tk.END)
            self.schedule_text.delete(1.0, tk.END)

            # Placeholders
            placeholder_analysis = (
                "\n" 
                "=====================================================\n"
                "CONTRACT ANALYSIS RESULTS WILL APPEAR HERE\n"
                "=====================================================\n\n"
                "Select a contract file and click 'Analyze Contract'\n"
                "to see comprehensive contract analysis results.\n"
            )

            placeholder_schedule = (
                "\n"
                "=====================================================\n"
                "PAYMENT SCHEDULE WILL APPEAR HERE\n"
                "=====================================================\n\n"
                "Payment schedule will be generated based on actual\n"
                "contract terms (no assumptions made).\n"
            )

            self.analysis_text.insert(tk.END, placeholder_analysis)
            self.schedule_text.insert(tk.END, placeholder_schedule)

            # Reset state variables
            self.current_analysis = None
            self.current_schedule = None
            self.export_button.config(state=tk.DISABLED)

            # Reset progress/timer/status
            self._stop_pulse()
            self._stop_timer()
            self._set_progress(0)
            self._set_status("Ready - No assumptions, analysis-driven only")

            # Show first tab
            self.notebook.select(0)

        except Exception as e:
            print(f"Error clearing results: {e}")

    def _display_results(self, analysis: Dict, payment_schedule: List[Dict]):
        """Display analysis results in GUI."""

        self.current_analysis = analysis
        self.current_schedule = payment_schedule

        # Display contract analysis
        self.analysis_text.delete(1.0, tk.END)

        if "error" in analysis:
            self.analysis_text.insert(tk.END, f"Analysis Error: {analysis['error']}\n")
        else:
            self.analysis_text.insert(tk.END, "CONTRACT ANALYSIS RESULTS\n")
            self.analysis_text.insert(tk.END, "=" * 50 + "\n\n")

            # Contract Information
            if "contract_info" in analysis:
                self.analysis_text.insert(tk.END, "CONTRACT INFORMATION:\n")
                contract_info = analysis["contract_info"]
                for key, value in contract_info.items():
                    formatted_key = key.replace('_', ' ').title()
                    formatted_value = value if value else "Not specified in contract"
                    self.analysis_text.insert(tk.END, f"  {formatted_key}: {formatted_value}\n")
                self.analysis_text.insert(tk.END, "\n")

            # Payment Terms
            if "payment_terms" in analysis:
                self.analysis_text.insert(tk.END, "PAYMENT TERMS:\n")
                payment_terms = analysis["payment_terms"]
                for key, value in payment_terms.items():
                    if key == "specific_payments":
                        continue  # Handle separately
                    formatted_key = key.replace('_', ' ').title()
                    formatted_value = value if value else "Not specified in contract"
                    self.analysis_text.insert(tk.END, f"  {formatted_key}: {formatted_value}\n")
                self.analysis_text.insert(tk.END, "\n")

            # Analysis Summary
            if "contract_analysis" in analysis:
                self.analysis_text.insert(tk.END, "CONTRACT ANALYSIS SUMMARY:\n")
                contract_analysis = analysis["contract_analysis"]
                for key, value in contract_analysis.items():
                    formatted_key = key.replace('_', ' ').title()
                    self.analysis_text.insert(tk.END, f"  {formatted_key}: {value}\n")

        # Display payment schedule
        self.schedule_text.delete(1.0, tk.END)

        if payment_schedule:
            self.schedule_text.insert(tk.END, "PAYMENT SCHEDULE (From Contract Analysis)\n")
            self.schedule_text.insert(tk.END, "=" * 50 + "\n\n")

            for i, payment in enumerate(payment_schedule, 1):
                self.schedule_text.insert(tk.END, f"Payment {i}:\n")
                self.schedule_text.insert(tk.END, f"  Amount: {payment.get('amount', 'Not specified')}\n")
                self.schedule_text.insert(tk.END, f"  Due Date: {payment.get('due_date', 'Not specified')}\n")
                self.schedule_text.insert(tk.END, f"  Description: {payment.get('description', 'Payment')}\n")
                self.schedule_text.insert(tk.END, f"  Source: {payment.get('source', 'Contract')}\n\n")
        else:
            self.schedule_text.insert(tk.END, "NO PAYMENT SCHEDULE SPECIFIED IN CONTRACT\n\n")
            self.schedule_text.insert(tk.END, "This contract does not contain explicit payment schedule information.\n")
            self.schedule_text.insert(tk.END, "The analyzer does not make assumptions about payment terms.\n\n")
            self.schedule_text.insert(tk.END, "To generate a payment schedule, the contract must explicitly specify:\n")
            self.schedule_text.insert(tk.END, "- Payment amounts\n")
            self.schedule_text.insert(tk.END, "- Payment dates or schedule\n")
            self.schedule_text.insert(tk.END, "- Payment conditions or triggers\n")

        # Re-enable buttons
        self.analyze_button.config(state=tk.NORMAL)
        self.export_button.config(state=tk.NORMAL)

    # ------------- Export -------------
    def export_results(self):
        """Export analysis results to Excel."""
        if not self.current_analysis:
            messagebox.showerror("Error", "No analysis results to export")
            return

        file_path = filedialog.asksaveasfilename(
            title="Save Analysis Results",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )

        if not file_path:
            return

        def do_export():
            try:
                self.root.after(0, lambda: (self._set_status("Exporting to Excel..."), self._set_progress(96)))
                self.analyzer.export_analysis_to_excel(
                    self.current_analysis,
                    self.current_schedule or [],
                    file_path
                )
                self.root.after(0, lambda: (self._set_progress(100), self._set_status(f"Exported to {file_path}")))
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Export failed: {str(e)}"))
                self.root.after(0, lambda: self._set_status("Export failed"))

        threading.Thread(target=do_export, daemon=True).start()

    # ------------- Run -------------
    def run(self):
        """Start the GUI application."""
        self.root.mainloop()


if __name__ == "__main__":
    app = ContractAnalyzerGUI()
    app.run()