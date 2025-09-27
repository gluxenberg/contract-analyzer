#!/usr/bin/env python3
"""
Contract Analyzer Core Library
Extracted from contract_analyzer_gui.py - Shared functionality between CLI and GUI versions
"""

import os
import sys
import json
import re
import tempfile
from pathlib import Path
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.differential import DifferentialStyle
import anthropic
from typing import Dict, List, Optional, Tuple

class HighPrecisionContractAnalyzer:
    def __init__(self, api_key: str):
        """Initialize the enhanced contract analyzer with Claude API key and precision patterns."""
        self.client = anthropic.Anthropic(api_key=api_key)

        # High-precision patterns for critical financial data
        self.financial_patterns = {
            'money_amounts': [
                r'\$\s*[\d,]+\.?\d*(?:\s*(?:USD|dollars?))?',
                r'(?:USD|dollars?)\s*[\d,]+\.?\d*',
                r'[\d,]+\.?\d*\s*(?:USD|dollars?)',
                r'(?:total|amount|value|fee|rate|cost|price|budget|cap|maximum|limit)(?:\s+(?:of|is|at|shall be|not to exceed))?\s*[:\$]?\s*[\d,]+\.?\d*',
            ],
            'hourly_rates': [
                r'\$\s*[\d,]+\.?\d*\s*(?:per|/|an)\s*hour',
                r'hourly\s+(?:rate|fee|charge)(?:\s+(?:of|is|at))?\s*[:\$]?\s*[\d,]+\.?\d*',
                r'[\d,]+\.?\d*\s*(?:per|/)\s*(?:hour|hr)',
            ],
            'contract_values': [
                r'(?:total|contract|project|agreement)\s+(?:value|amount|price|cost|fee)(?:\s+(?:of|is|shall be|not to exceed))?\s*[:\$]?\s*[\d,]+\.?\d*',
                r'(?:maximum|cap|limit|ceiling)(?:\s+(?:of|is|at))?\s*[:\$]?\s*[\d,]+\.?\d*',
                r'budget(?:\s+(?:of|is|shall be|not to exceed))?\s*[:\$]?\s*[\d,]+\.?\d*',
            ],
            'dates': [
                r'\b\d{1,2}[/-]\d{1,2}[/-]\d{4}\b',
                r'\b\d{4}[/-]\d{2}[/-]\d{2}\b',
                r'\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},?\s+\d{4}\b',
                r'\b\d{1,2}\s+(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}\b',
            ],
            'time_periods': [
                r'\b\d+\s*(?:hours?|hrs?)\b',
                r'\b\d+\s*(?:days?)\b',
                r'\b\d+\s*(?:weeks?)\b',
                r'\b\d+\s*(?:months?)\b',
                r'maximum.*?\d+.*?(?:hours?|days?)',
                r'(?:daily|weekly|monthly)\s+(?:limit|maximum|cap).*?\d+',
            ],
            'payment_terms': [
                r'net\s+\d+(?:\s+days?)?',
                r'within\s+\d+\s+days?',
                r'payment\s+due.*?\d+\s+days?',
                r'\d+\s+days?\s+(?:after|from|following)',
            ]
        }

    def parse_currency_to_number(self, currency_string: str) -> Optional[float]:
        """Convert currency string to float number for Excel calculations."""
        if not currency_string or currency_string in ['null', 'N/A', 'Not found', '']:
            return None

        temp_str = str(currency_string).strip()
        if temp_str.startswith(','):
            temp_str = '0' + temp_str
        if temp_str.startswith('.'):
            temp_str = '0' + temp_str
        cleaned = re.sub(r'[^\d.-]', '', temp_str)
        try:
            return float(cleaned) if cleaned else None
        except ValueError:
            return None

    def parse_hours_to_number(self, hours_string: str) -> Optional[float]:
        """Convert hours string to float number."""
        if not hours_string or hours_string in ['null', 'N/A', 'Not found', '']:
            return None

        temp_str = str(hours_string).strip()
        if temp_str.startswith(','):
            temp_str = '0' + temp_str
        if temp_str.startswith('.'):
            temp_str = '0' + temp_str
        cleaned = re.sub(r'[^\d.-]', '', temp_str)
        try:
            return float(cleaned) if cleaned else None
        except ValueError:
            return None

    def read_contract_file(self, file_path: str) -> str:
        """Read contract file and return content as string."""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Contract file not found: {file_path}")

        if path.suffix.lower() == '.pdf':
            try:
                import PyPDF2
                with open(file_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    text = ""
                    for page in pdf_reader.pages:
                        text += page.extract_text()
                    return text
            except ImportError:
                raise ImportError("PyPDF2 required for PDF files. Install with: pip install PyPDF2")

        elif path.suffix.lower() in ['.txt', '.docx', '.doc']:
            if path.suffix.lower() == '.docx':
                try:
                    from docx import Document
                    doc = Document(file_path)
                    text = ""
                    for paragraph in doc.paragraphs:
                        text += paragraph.text + "\n"
                    return text
                except ImportError:
                    raise ImportError("python-docx required for Word files. Install with: pip install python-docx")
            else:
                # Plain text
                try:
                    with open(file_path, 'r', encoding='utf-8') as file:
                        return file.read()
                except UnicodeDecodeError:
                    with open(file_path, 'r', encoding='latin-1') as file:
                        return file.read()
        else:
            raise ValueError(f"Unsupported file type: {path.suffix}")

    def extract_financial_data_regex(self, contract_text: str) -> Dict:
        """PASS 1: High-precision regex extraction for critical financial data."""
        results = {}

        for category, patterns in self.financial_patterns.items():
            matches = []
            for pattern in patterns:
                found = re.findall(pattern, contract_text, re.IGNORECASE | re.MULTILINE)
                for match in found:
                    if match not in [m[0] for m in matches]:  # Avoid duplicates
                        matches.append((match, len(match)))  # Include length for scoring

            # Sort by length (longer matches often more specific)
            matches.sort(key=lambda x: x[1], reverse=True)
            results[category] = matches[:10]  # Keep top 10 matches

        return results

    def claude_direct_extraction(self, contract_text: str) -> Dict:
        """PASS 2: Direct Claude extraction with focused prompts."""
        claude_prompt = f"""
        Extract ONLY these specific financial data points from this contract.
        Be extremely precise - include ONLY information that is explicitly stated:

        1. TOTAL CONTRACT VALUE (the main contract amount)
        2. HOURLY RATE (if time & materials)
        3. CONTRACT START DATE
        4. CONTRACT END DATE
        5. PAYMENT TERMS (Net 30, etc.)
        6. MAXIMUM HOURS (if specified)
        7. DAILY HOUR LIMIT (if specified)

        Return as JSON:
        {{
            "total_contract_value": "exact amount or null",
            "hourly_rate": "exact rate or null",
            "start_date": "YYYY-MM-DD or null",
            "end_date": "YYYY-MM-DD or null",
            "payment_terms": "exact terms or null",
            "maximum_hours": "number or null",
            "daily_hour_limit": "number or null"
        }}

        Contract: {contract_text}
        """

        try:
            claude_result = self.client.messages.create(
                model="claude-3-haiku-20240307",
                max_tokens=1000,
                messages=[{"role": "user", "content": claude_prompt}]
            )

            claude_text = claude_result.content[0].text
            claude_data = json.loads(claude_text)
            return claude_data

        except Exception as e:
            return {"error": f"Claude extraction failed: {str(e)}"}

    def validate_and_merge_data(self, regex_results: Dict, claude_data: Dict, contract_text: str) -> Dict:
        """PASS 3: Validate and merge results from both methods."""
        validation_prompt = f"""
        You are validating financial data extracted by two different methods.
        Provide the MOST ACCURATE data by comparing both sources:

        REGEX EXTRACTED DATA:
        Money amounts: {[m[0] for m in regex_results.get('money_amounts', [])[:5]]}
        Hourly rates: {[m[0] for m in regex_results.get('hourly_rates', [])]}
        Dates: {[m[0] for m in regex_results.get('dates', [])[:10]]}
        Payment terms: {[m[0] for m in regex_results.get('payment_terms', [])]}

        CLAUDE EXTRACTED DATA:
        {claude_data}

                Return ONLY valid JSON in this exact format (no other text):
        {{
            "total_contract_value": "actual_value_or_null",
            "hourly_rate": "actual_value_or_null",
            "start_date": "YYYY-MM-DD_or_null",
            "end_date": "YYYY-MM-DD_or_null",
            "payment_terms": "actual_terms_or_null",
            "maximum_hours": "number_or_null",
            "daily_hour_limit": "number_or_null",
            "confidence_scores": {{
                "total_contract_value": "HIGH",
                "hourly_rate": "MEDIUM",
                "start_date": "HIGH",
                "end_date": "HIGH",
                "payment_terms": "HIGH",
                "maximum_hours": "LOW",
                "daily_hour_limit": "LOW"
            }},
            "validation_notes": "Brief summary of data quality"
        }}
        """

        try:
            validation_result = self.client.messages.create(
                model="claude-3-haiku-20240307",
                max_tokens=1500,
                messages=[{"role": "user", "content": validation_prompt}]
            )

            validation_text = validation_result.content[0].text
            validated_data = json.loads(validation_text)
            return validated_data

        except Exception as e:
            # Fallback to Claude data if validation fails
            fallback_data = claude_data.copy()
            fallback_data["validation_notes"] = f"Validation failed: {str(e)}"
            return fallback_data

    def cross_validate_financial_data(self, contract_text: str) -> Dict:
        """Main three-pass validation method."""
        # Pass 1: Regex extraction
        regex_results = self.extract_financial_data_regex(contract_text)

        # Pass 2: Claude direct extraction
        claude_data = self.claude_direct_extraction(contract_text)

        # Pass 3: Validation and merging
        validated_data = self.validate_and_merge_data(regex_results, claude_data, contract_text)
        return validated_data

    def analyze_contract_comprehensive(self, contract_text: str) -> Dict:
        """Send contract to Claude API for comprehensive financial analysis."""
        enhanced_analysis_prompt = """
        You are a Certified Public Accountant analyzing this contract for complete financial tracking requirements. Extract ALL financial information including:

        BASIC CONTRACT INFORMATION (EXTRACT FIRST):
        - Client name (party receiving services/goods)
        - Vendor/Contractor name (party providing services/goods)
        - Contract number or ID
        - Contract effective dates
        - Contract title or project name

        Then extract ALL financial information including:

        PRIMARY PAYMENT STRUCTURE:
        - Contract type (fixed payments, time & materials, milestone-based, etc.)
        - Total contract value and any caps/maximums
        - Payment rates (hourly, daily, fixed amounts)
        - Payment frequency and timing requirements

        PAYMENT SCHEDULE DETAILS:
        - Specific due dates for payments OR invoice submission requirements
        - Invoice approval processes and timeframes
        - Expected payment timing after invoice submission
        - Any advance payments or retainer requirements

        EXPENSE AND REIMBURSEMENT TRACKING:
        - Travel expense policies and limits
        - Meal and incidental expense allowances
        - Equipment or material cost provisions
        - Any expense pre-approval requirements
        - Separate expense tracking from main contract value

        COMPLIANCE AND DOCUMENTATION REQUIREMENTS:
        - Required tax forms (W-9, 1099, etc.)
        - Invoice content requirements (daily breakdowns, certifications, etc.)
        - Time tracking limitations (daily/weekly maximums)
        - Reporting deadlines and deliverable schedules
        - Required signatures or approvals

        BUDGET MONITORING ELEMENTS:
        - Maximum hours, days, or other quantity limits
        - Running total calculations needed
        - Budget utilization warnings or thresholds
        - Contract period start and end dates
        - Any renewal or extension provisions

        RISK FACTORS:
        - Payment conditions or performance requirements
        - Penalty clauses or withholding provisions
        - Termination clauses affecting payment
        - Currency or rate change provisions

        For time & materials contracts specifically include:
        - Hourly rate breakdown by role/activity
        - Maximum billable hours per period
        - Overtime or premium rate conditions
        - Minimum/maximum monthly billing requirements

        Create a comprehensive payment tracking structure that includes:
        1. All scheduled payments with dates and amounts
        2. Invoice submission timeline and requirements
        3. Expense reimbursement tracking separate from main payments
        4. Compliance documentation deadlines
        5. Budget utilization monitoring with warnings
        6. Expected payment dates based on contract terms

        If no specific payment dates exist, create a logical payment tracking framework based on the contract's invoicing and payment cycle requirements.

        Contract text:
        """

        try:
            message = self.client.messages.create(
                model="claude-3-haiku-20240307",
                max_tokens=4000,
                messages=[{
                    "role": "user",
                    "content": enhanced_analysis_prompt + contract_text
                }]
            )
            return {"analysis": message.content[0].text, "success": True}
        except Exception as e:
            return {"analysis": f"Error analyzing contract: {str(e)}", "success": False}

    def extract_comprehensive_data(self, analysis: str) -> Tuple[Dict, List[Dict], Dict]:
        """Extract comprehensive structured data from Claude's analysis."""
        extraction_prompt = f"""
        Based on this contract analysis, extract the comprehensive information in JSON format:

        Return a JSON object with:

        1. "contract_info": {{
            "client": "client name",
            "vendor": "vendor name",
            "contract_id": "contract number/id",
            "contract_type": "fixed payments/time and materials/milestone/other",
            "total_value": "total contract value",
            "start_date": "YYYY-MM-DD",
            "end_date": "YYYY-MM-DD",
            "payment_terms": "payment terms description",
            "hourly_rate": "hourly rate if applicable",
            "max_hours": "maximum hours if applicable",
            "max_daily_hours": "daily hour limit if applicable",
            "invoice_frequency": "monthly/weekly/other",
            "payment_timeline": "net 30/45 days or description"
        }}

        2. "payment_schedule": [
            {{
                "due_date": "YYYY-MM-DD or null",
                "description": "payment description",
                "amount": "payment amount or null",
                "invoice_submission_due": "YYYY-MM-DD or null",
                "expected_payment_date": "YYYY-MM-DD or null",
                "notes": "any additional notes"
            }}
        ]

        3. "tracking_requirements": {{
            "expense_tracking": {{
                "travel_expenses": "yes/no",
                "travel_policy": "description",
                "meal_allowance": "amount or policy",
                "equipment_costs": "yes/no",
                "pre_approval_required": "yes/no"
            }},
            "compliance": {{
                "w9_required": "yes/no",
                "invoice_requirements": "description",
                "time_breakdown_required": "yes/no",
                "certifications_required": "description",
                "reporting_deadlines": "description"
            }},
            "budget_monitoring": {{
                "hour_tracking": "yes/no",
                "budget_caps": "description",
                "warning_thresholds": "percentage or amount",
                "utilization_tracking": "yes/no"
            }}
        }}

        If information is not available, use null or "not specified".

        Analysis to process:
        {analysis}
        """

        try:
            message = self.client.messages.create(
                model="claude-3-haiku-20240307",
                max_tokens=2000,
                messages=[{
                    "role": "user",
                    "content": extraction_prompt
                }]
            )

            response_text = message.content[0].text.strip()

            # Try to clean up the JSON response
            if "```json" in response_text:
                # Extract JSON from code blocks
                start = response_text.find("```json") + 7
                end = response_text.find("```", start)
                if end != -1:
                    response_text = response_text[start:end].strip()

            result = json.loads(response_text)

            contract_info = result.get("contract_info", {})
            payment_schedule = result.get("payment_schedule", [])
            tracking_requirements = result.get("tracking_requirements", {})

            return contract_info, payment_schedule, tracking_requirements

        except json.JSONDecodeError as e:
            # Fallback: Try to extract basic info from the analysis text
            return self._extract_fallback_data(analysis, str(e))
        except Exception as e:
            # Return structures with fallback data
            return self._extract_fallback_data(analysis, str(e))

    def _extract_fallback_data(self, analysis_text: str, error_msg: str) -> Tuple[Dict, List[Dict], Dict]:
        """Fallback method to extract basic contract info from analysis text"""

        contract_info = {
            "client": self._extract_from_text(analysis_text, ["CLIENT:", "Client:", "client name", "receiving services"]),
            "vendor": self._extract_from_text(analysis_text, ["CONTRACTOR:", "Vendor:", "vendor name", "providing services"]),
            "contract_id": self._extract_from_text(analysis_text, ["Contract ID", "Agreement", "contract number"]),
            "contract_type": self._extract_contract_type(analysis_text),
            "total_value": "unknown",
            "start_date": None,
            "end_date": None,
            "payment_terms": "unknown",
            "hourly_rate": "unknown",
            "max_hours": None,
            "max_daily_hours": None,
            "invoice_frequency": "unknown",
            "payment_timeline": "unknown"
        }

        # Basic payment schedule
        payment_schedule = [{
            "due_date": None,
            "description": "Monthly invoice payment",
            "amount": None,
            "invoice_submission_due": None,
            "expected_payment_date": None,
            "notes": f"Extraction failed: {error_msg}"
        }]

        # Basic tracking requirements
        tracking_requirements = {
            "expense_tracking": {
                "travel_expenses": "unknown",
                "travel_policy": "unknown",
                "meal_allowance": "unknown",
                "equipment_costs": "unknown",
                "pre_approval_required": "unknown"
            },
            "compliance": {
                "w9_required": "unknown",
                "invoice_requirements": "unknown",
                "time_breakdown_required": "unknown",
                "certifications_required": "unknown",
                "reporting_deadlines": "unknown"
            },
            "budget_monitoring": {
                "hour_tracking": "unknown",
                "budget_caps": "unknown",
                "warning_thresholds": "unknown",
                "utilization_tracking": "unknown"
            }
        }

        return contract_info, payment_schedule, tracking_requirements

    def _extract_from_text(self, text: str, search_terms: List[str]) -> str:
        """Extract information from text using search terms"""
        text_lower = text.lower()

        for term in search_terms:
            term_lower = term.lower()
            if term_lower in text_lower:
                # Find the line containing the term
                lines = text.split('\n')
                for line in lines:
                    if term_lower in line.lower():
                        # Extract the part after the term
                        parts = line.split(':')
                        if len(parts) > 1:
                            result = parts[1].strip()
                            if result and len(result) > 2:
                                return result[:50]  # Limit length

        return "Not specified"

    def _extract_contract_type(self, text: str) -> str:
        """Extract contract type from analysis text"""
        text_lower = text.lower()

        if any(term in text_lower for term in ["time and materials", "hourly", "per hour"]):
            return "Time and Materials"
        elif any(term in text_lower for term in ["fixed price", "lump sum", "total contract"]):
            return "Fixed Price"
        elif any(term in text_lower for term in ["milestone", "deliverable"]):
            return "Milestone-based"
        else:
            return "Unknown"

    def create_comprehensive_spreadsheet(self, contract_info: Dict, payment_schedule: List[Dict],
                                       tracking_requirements: Dict, validated_data: Dict,
                                       analysis_text: str, output_file: str):
        """Create comprehensive Excel spreadsheet with all tracking sheets."""

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 1. Contract Summary Sheet
        summary_ws = wb.create_sheet("Contract Summary")

        summary_data = [
            ["CONTRACT OVERVIEW", ""],
            ["Client", contract_info.get('client', 'N/A')],
            ["Vendor/Contractor", contract_info.get('vendor', 'N/A')],
            ["Contract ID", contract_info.get('contract_id', 'N/A')],
            ["Contract Type", contract_info.get('contract_type', 'N/A')],
            ["Total Value", contract_info.get('total_value', 'N/A')],
            ["Start Date", contract_info.get('start_date', 'N/A')],
            ["End Date", contract_info.get('end_date', 'N/A')],
            ["", ""],
            ["HIGH-PRECISION FINANCIAL DATA", ""],
            ["Total Contract Value", validated_data.get('total_contract_value', 'N/A')],
            ["Hourly Rate", validated_data.get('hourly_rate', 'N/A')],
            ["Payment Terms", validated_data.get('payment_terms', 'N/A')],
            ["Maximum Hours", validated_data.get('maximum_hours', 'N/A')],
            ["Daily Hour Limit", validated_data.get('daily_hour_limit', 'N/A')],
            ["", ""],
            ["CONFIDENCE SCORES", ""],
        ]

        # Add confidence scores if available
        confidence_scores = validated_data.get('confidence_scores', {})
        for field, score in confidence_scores.items():
            summary_data.append([field.replace('_', ' ').title(), score])

        for row_idx, (label, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row_idx, column=1, value=label)
            summary_ws.cell(row=row_idx, column=2, value=value)

            if label in ["CONTRACT OVERVIEW", "HIGH-PRECISION FINANCIAL DATA", "CONFIDENCE SCORES"]:
                summary_ws.cell(row=row_idx, column=1).font = header_font
                summary_ws.cell(row=row_idx, column=1).fill = header_fill

        summary_ws.column_dimensions['A'].width = 30
        summary_ws.column_dimensions['B'].width = 40

        # 2. Payment Tracking Sheet
        payment_ws = wb.create_sheet("Payment Tracking")
        payment_headers = ["Due Date", "Description", "Amount", "Invoice Due", "Expected Payment", "Status", "Notes"]

        for col_idx, header in enumerate(payment_headers, 1):
            cell = payment_ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, payment in enumerate(payment_schedule, 2):
            payment_ws.cell(row=row_idx, column=1, value=payment.get('due_date', ''))
            payment_ws.cell(row=row_idx, column=2, value=payment.get('description', ''))
            payment_ws.cell(row=row_idx, column=3, value=payment.get('amount', ''))
            payment_ws.cell(row=row_idx, column=4, value=payment.get('invoice_submission_due', ''))
            payment_ws.cell(row=row_idx, column=5, value=payment.get('expected_payment_date', ''))
            payment_ws.cell(row=row_idx, column=6, value='Pending')
            payment_ws.cell(row=row_idx, column=7, value=payment.get('notes', ''))

        # Auto-fit columns
        for col in payment_ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            payment_ws.column_dimensions[column].width = adjusted_width

        # 3. Expense Tracking Sheet
        expense_ws = wb.create_sheet("Expense Tracking")
        expense_headers = ["Date", "Category", "Description", "Amount", "Receipt", "Approved", "Reimbursed", "Notes"]

        for col_idx, header in enumerate(expense_headers, 1):
            cell = expense_ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Add expense policy info
        expense_tracking = tracking_requirements.get('expense_tracking', {})
        expense_ws.cell(row=3, column=1, value="Travel Expenses Allowed:")
        expense_ws.cell(row=3, column=2, value=expense_tracking.get('travel_expenses', 'Not specified'))
        expense_ws.cell(row=4, column=1, value="Travel Policy:")
        expense_ws.cell(row=4, column=2, value=expense_tracking.get('travel_policy', 'Not specified'))

        for col in expense_ws.columns:
            max_length = 15
            column = col[0].column_letter
            expense_ws.column_dimensions[column].width = max_length

        # 4. Compliance Checklist Sheet
        compliance_ws = wb.create_sheet("Compliance Checklist")
        compliance_headers = ["Requirement", "Status", "Due Date", "Completed Date", "Notes"]

        for col_idx, header in enumerate(compliance_headers, 1):
            cell = compliance_ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Add compliance requirements
        compliance = tracking_requirements.get('compliance', {})
        compliance_items = [
            ("W-9 Form Required", compliance.get('w9_required', 'Not specified')),
            ("Invoice Requirements", compliance.get('invoice_requirements', 'Not specified')),
            ("Time Breakdown Required", compliance.get('time_breakdown_required', 'Not specified')),
            ("Certifications Required", compliance.get('certifications_required', 'Not specified')),
            ("Reporting Deadlines", compliance.get('reporting_deadlines', 'Not specified'))
        ]

        for row_idx, (requirement, status) in enumerate(compliance_items, 2):
            compliance_ws.cell(row=row_idx, column=1, value=requirement)
            compliance_ws.cell(row=row_idx, column=2, value=status)
            compliance_ws.cell(row=row_idx, column=3, value="")  # Due date
            compliance_ws.cell(row=row_idx, column=4, value="")  # Completed date
            compliance_ws.cell(row=row_idx, column=5, value="")  # Notes

        for col in compliance_ws.columns:
            column = col[0].column_letter
            compliance_ws.column_dimensions[column].width = 20

        # 5. Budget Monitor Sheet
        budget_ws = wb.create_sheet("Budget Monitor")
        budget_headers = ["Period", "Hours Used", "Amount Spent", "% Budget Used", "Remaining Budget", "Status"]

        for col_idx, header in enumerate(budget_headers, 1):
            cell = budget_ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Add budget monitoring info
        budget_monitoring = tracking_requirements.get('budget_monitoring', {})
        budget_ws.cell(row=3, column=1, value="Budget Caps:")
        budget_ws.cell(row=3, column=2, value=budget_monitoring.get('budget_caps', 'Not specified'))
        budget_ws.cell(row=4, column=1, value="Warning Thresholds:")
        budget_ws.cell(row=4, column=2, value=budget_monitoring.get('warning_thresholds', 'Not specified'))

        for col in budget_ws.columns:
            column = col[0].column_letter
            budget_ws.column_dimensions[column].width = 18

        # 6. Validation Details Sheet
        validation_ws = wb.create_sheet("Validation Details")

        validation_data = [
            ["THREE-PASS VALIDATION SYSTEM", ""],
            ["", ""],
            ["VALIDATION METHODOLOGY", ""],
            ["Pass 1", "High-precision regex pattern extraction"],
            ["Pass 2", "Direct Claude AI extraction with focused prompts"],
            ["Pass 3", "Cross-validation and confidence scoring"],
            ["", ""],
            ["CONFIDENCE LEVELS", ""],
            ["HIGH", "Data confirmed by multiple sources"],
            ["MEDIUM", "Data found but requires verification"],
            ["LOW", "Data uncertain or not clearly stated"],
            ["", ""],
            ["VALIDATION NOTES", ""],
            ["Summary", validated_data.get('validation_notes', 'No validation notes available')]
        ]

        for row_idx, (label, value) in enumerate(validation_data, 1):
            validation_ws.cell(row=row_idx, column=1, value=label)
            validation_ws.cell(row=row_idx, column=2, value=value)

            if label in ["THREE-PASS VALIDATION SYSTEM", "VALIDATION METHODOLOGY", "CONFIDENCE LEVELS", "VALIDATION NOTES"]:
                validation_ws.cell(row=row_idx, column=1).font = header_font
                validation_ws.cell(row=row_idx, column=1).fill = header_fill

        validation_ws.column_dimensions['A'].width = 25
        validation_ws.column_dimensions['B'].width = 50

        # 7. Full Analysis Sheet
        analysis_ws = wb.create_sheet("Full Analysis")
        analysis_ws.cell(row=1, column=1, value="COMPLETE CONTRACT ANALYSIS")
        analysis_ws.cell(row=1, column=1).font = header_font
        analysis_ws.cell(row=1, column=1).fill = header_fill

        # Split analysis text into lines and add to sheet
        analysis_lines = analysis_text.split('\n')
        for row_idx, line in enumerate(analysis_lines, 3):
            analysis_ws.cell(row=row_idx, column=1, value=line)

        analysis_ws.column_dimensions['A'].width = 100

        wb.save(output_file)